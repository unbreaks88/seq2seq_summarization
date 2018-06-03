import csv

import openpyxl
import os

title_col = 'E'
category_col = 'F'
keyword_col = 'O'
feature_col = 'P'
contents = 'Q'
except_col = 'S'

col_list = [title_col, category_col, keyword_col, feature_col, contents]


def get_file_list(dir):
    file_list = []
    filenames = os.listdir(dir)
    for filename in filenames:
        file_list.append(os.path.join(dir, filename))
    print('get file list')
    print(file_list)
    return file_list


def load_excel(filename):
    print("load_excel: ", filename)
    workbook = openpyxl.load_workbook(filename=filename)
    # sheet = workbook.get_sheet_by_name('sheet')
    sheet = workbook['sheet']
    print('max_row: ', sheet.max_row)
    print('max_column', sheet.max_column)

    return sheet


def check_valid_data(value):
    if value is None:
        return False
    else:
        return True


def parse_data(sheet):
    documents = []
    for i in range(2, sheet.max_row):
        # if check_valid_data(sheet[except_col + str(i)].value):
        title_val = sheet[title_col + str(i)].value
        category_val = sheet[category_col + str(i)].value
        # keyword_val = sheet[keyword_col + str(i)].value
        feature_val = sheet[feature_col + str(i)].value
        contents_val = sheet[contents + str(i)].value
        # document = [title_val.strip(), category_val.strip(), keyword_val.strip(), feature_val.strip(),
        #             contents_val.strip()]
        document = [title_val.strip(), category_val.strip(), contents_val.strip()]
        documents.append(document)

    print('document_length {}: '.format(len(documents)))
    return documents


def write_csv(filename, documents):
    with open(filename, 'w', encoding='utf-8', newline='') as csvfile:
        writer = csv.writer(csvfile, delimiter='\t')
        for document in documents:
            writer.writerow(document)


if __name__ == '__main__':
    dir_name = '.data/'
    file_list = get_file_list(dir_name)

    for idx, file in enumerate(file_list):
        if not file.startswith('.'):
            sheet = load_excel(file)
            write_csv('/Users/unbreaks/Desktop/news/news_' + str(idx) + '.csv', parse_data(sheet))
